import io
import uuid
import os
from functools import wraps
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
from flask import session, redirect, url_for, flash, request, send_file
from werkzeug.utils import secure_filename


def allowed_file(filename, allowed_extensions):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in allowed_extensions


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("请先登录", "warning")
            return redirect(url_for("auth.login"))
        return f(*args, **kwargs)
    return decorated


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if "user_id" not in session:
                flash("请先登录", "warning")
                return redirect(url_for("auth.login"))
            if session.get("role") not in roles:
                flash("权限不足", "danger")
                return redirect(url_for("main.index"))
            return f(*args, **kwargs)
        return decorated
    return decorator


def get_manager_building_id():
    from models import User, db
    if session.get("role") == "dorm_manager":
        user = db.session.get(User, session["user_id"])
        return user.building_id if user else None
    return None


def log_operation(action, target_type=None, target_id=None):
    from models import db, OperationLog
    log = OperationLog(
        user_id=session["user_id"],
        action=action,
        target_type=target_type,
        target_id=target_id,
        ip_address=request.remote_addr,
    )
    db.session.add(log)
    db.session.commit()


def export_excel(filename, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename
    header_font = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{filename}.xlsx",
    )


def save_upload_file(file, upload_folder):
    ext = file.filename.rsplit(".", 1)[1].lower()
    filename = f"{uuid.uuid4().hex}.{ext}"
    file.save(os.path.join(upload_folder, filename))
    return filename
