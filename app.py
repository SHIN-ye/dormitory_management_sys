import os
import re
import io
import uuid
from functools import wraps
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, send_from_directory
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from config import Config
from sqlalchemy import text
from models import db, Building, Room, Student, Accommodation, Repair, Visitor, Fee, User, Announcement, CheckoutRequest, TransferRequest, OperationLog

app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in app.config["ALLOWED_EXTENSIONS"]


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("请先登录", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if "user_id" not in session:
                flash("请先登录", "warning")
                return redirect(url_for("login"))
            if session.get("role") not in roles:
                flash("权限不足", "danger")
                return redirect(url_for("index"))
            return f(*args, **kwargs)
        return decorated
    return decorator


def get_manager_building_id():
    """返回 dorm_manager 绑定的宿舍楼 ID，admin 返回 None 表示不限范围"""
    if session.get("role") == "dorm_manager":
        user = db.session.get(User, session["user_id"])
        return user.building_id if user else None
    return None


def log_operation(action, target_type=None, target_id=None):
    """记录操作日志，写入独立事务"""
    log = OperationLog(
        user_id=session["user_id"],
        action=action,
        target_type=target_type,
        target_id=target_id,
        ip_address=request.remote_addr,
    )
    db.session.add(log)
    db.session.commit()


def export_excel(filename, headers, rows):
    """生成 Excel 文件并通过 send_file 返回下载响应"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename
    header_font = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{filename}.xlsx",
    )


# ============================================================
# 登录 / 登出
# ============================================================

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            session["user_id"] = user.id
            session["username"] = user.username
            session["role"] = user.role
            flash("登录成功", "success")
            return redirect(url_for("index"))
        flash("用户名或密码错误", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("已退出登录", "info")
    return redirect(url_for("login"))


# ============================================================
# 首页仪表盘
# ============================================================

@app.route("/")
@login_required
def index():
    latest_announcements = Announcement.query.order_by(Announcement.created_at.desc()).limit(3).all()

    # 学生端：个人概览
    if session.get("role") == "student":
        student = Student.query.filter_by(user_id=session["user_id"]).first()
        my_accommodation = None
        roommates = []
        my_repairs = []
        my_fees = []
        if student:
            my_accommodation = Accommodation.query.filter_by(
                student_id=student.id, status="入住"
            ).first()
            if my_accommodation:
                # 室友
                roommates = Accommodation.query.options(
                    db.joinedload(Accommodation.student)
                ).filter(
                    Accommodation.room_id == my_accommodation.room_id,
                    Accommodation.status == "入住",
                    Accommodation.student_id != student.id,
                ).all()
            my_repairs = Repair.query.filter_by(student_id=student.id).order_by(Repair.report_date.desc()).limit(5).all()
            my_fees = Fee.query.filter_by(student_id=student.id).order_by(Fee.due_date.desc()).limit(5).all()
        return render_template("index.html",
            announcements=latest_announcements,
            is_student=True,
            student=student,
            my_accommodation=my_accommodation,
            roommates=roommates,
            my_repairs=my_repairs,
            my_fees=my_fees,
        )

    # 管理员/宿管端：全局统计
    manager_bid = get_manager_building_id()
    stats = {
        "buildings": Building.query.count() if not manager_bid else 1,
        "rooms": Room.query.filter_by(building_id=manager_bid).count() if manager_bid else Room.query.count(),
        "students": Student.query.count(),
        "occupied": Accommodation.query.join(Room).filter(Accommodation.status == "入住", Room.building_id == manager_bid).count() if manager_bid else Accommodation.query.filter_by(status="入住").count(),
        "pending_repairs": Repair.query.join(Room).filter(Repair.status == "待处理", Room.building_id == manager_bid).count() if manager_bid else Repair.query.filter_by(status="待处理").count(),
        "unpaid_fees": Fee.query.join(Room).filter(Fee.status == "未缴", Room.building_id == manager_bid).count() if manager_bid else Fee.query.filter_by(status="未缴").count(),
    }
    # 各宿舍楼入住率（调用 MySQL 函数）
    buildings = Building.query.all() if not manager_bid else Building.query.filter_by(id=manager_bid).all()
    occupancy_rates = {}
    for b in buildings:
        row = db.session.execute(text("SELECT fn_occupancy_rate(:bid)"), {"bid": b.id}).fetchone()
        occupancy_rates[b.id] = float(row[0]) if row else 0.0
    recent_repairs = (Repair.query.join(Room).filter(Room.building_id == manager_bid).order_by(Repair.report_date.desc()).limit(5).all()
        if manager_bid else Repair.query.order_by(Repair.report_date.desc()).limit(5).all())
    return render_template("index.html", stats=stats, recent_repairs=recent_repairs,
        announcements=latest_announcements, is_student=False,
        buildings=buildings, occupancy_rates=occupancy_rates)


# ============================================================
# 宿舍楼管理
# ============================================================

@app.route("/buildings")
@role_required("admin")
def building_list():
    search = request.args.get("search", "").strip()
    page = request.args.get("page", 1, type=int)
    q = Building.query
    if search:
        q = q.filter(db.or_(Building.name.contains(search), Building.address.contains(search), Building.manager.contains(search)))
    pagination = q.order_by(Building.id).paginate(page=page, per_page=15)
    return render_template("buildings.html", buildings=pagination.items, pagination=pagination, search=search)


@app.route("/buildings/export")
@role_required("admin")
def building_export():
    search = request.args.get("search", "").strip()
    q = Building.query
    if search:
        q = q.filter(db.or_(Building.name.contains(search), Building.address.contains(search), Building.manager.contains(search)))
    buildings = q.order_by(Building.id).all()
    return export_excel(
        "宿舍楼列表",
        ["名称", "层数", "地址", "管理员"],
        [[b.name, f"{b.floors}层", b.address or "-", b.manager or "-"] for b in buildings],
    )


@app.route("/building/add", methods=["GET", "POST"])
@role_required("admin")
def building_add():
    if request.method == "POST":
        floors = int(request.form.get("floors", 6))
        if floors < 1:
            flash("楼层数至少为1", "warning")
            return render_template("building_form.html", building=None)
        b = Building(
            name=request.form["name"],
            floors=floors,
            address=request.form.get("address", ""),
            manager=request.form.get("manager", ""),
        )
        db.session.add(b)
        db.session.commit()
        log_operation(f"添加宿舍楼「{b.name}」", "building", b.id)
        flash("添加成功", "success")
        return redirect(url_for("building_list"))
    return render_template("building_form.html", building=None)


@app.route("/building/<int:bid>/edit", methods=["GET", "POST"])
@role_required("admin")
def building_edit(bid):
    b = Building.query.get_or_404(bid)
    if request.method == "POST":
        floors = int(request.form.get("floors", 6))
        if floors < 1:
            flash("楼层数至少为1", "warning")
            return render_template("building_form.html", building=b)
        b.name = request.form["name"]
        b.floors = floors
        b.address = request.form.get("address", "")
        b.manager = request.form.get("manager", "")
        db.session.commit()
        log_operation(f"修改宿舍楼「{b.name}」", "building", b.id)
        flash("修改成功", "success")
        return redirect(url_for("building_list"))
    return render_template("building_form.html", building=b)


@app.route("/building/<int:bid>/delete", methods=["POST"])
@role_required("admin")
def building_delete(bid):
    b = Building.query.get_or_404(bid)
    name = b.name
    db.session.delete(b)
    db.session.commit()
    log_operation(f"删除宿舍楼「{name}」", "building", bid)
    flash("删除成功", "success")
    return redirect(url_for("building_list"))


# ============================================================
# 房间管理
# ============================================================

@app.route("/rooms")
@role_required("admin")
def room_list():
    search = request.args.get("search", "").strip()
    bid = request.args.get("building_id", type=int)
    page = request.args.get("page", 1, type=int)
    q = Room.query
    if bid:
        q = q.filter_by(building_id=bid)
    if search:
        q = q.filter(Room.room_number.contains(search))
    pagination = q.order_by(Room.building_id, Room.room_number).paginate(page=page, per_page=15)
    buildings = Building.query.all()
    return render_template("rooms.html", rooms=pagination.items, buildings=buildings, current_building=bid, pagination=pagination, search=search)


@app.route("/rooms/export")
@role_required("admin")
def room_export():
    search = request.args.get("search", "").strip()
    bid = request.args.get("building_id", type=int)
    q = Room.query
    if bid:
        q = q.filter_by(building_id=bid)
    if search:
        q = q.filter(Room.room_number.contains(search))
    rooms = q.order_by(Room.building_id, Room.room_number).all()
    return export_excel(
        "房间列表",
        ["宿舍楼", "房间号", "类型", "容量", "已住", "价格(元/年)"],
        [[r.building.name, r.room_number, r.room_type, r.capacity, r.occupied, r.price] for r in rooms],
    )


@app.route("/room/add", methods=["GET", "POST"])
@role_required("admin")
def room_add():
    if request.method == "POST":
        capacity = int(request.form.get("capacity", 4))
        if capacity < 1:
            flash("房间容量至少为1", "warning")
            return render_template("room_form.html", room=None, buildings=Building.query.all())
        r = Room(
            room_number=request.form["room_number"],
            building_id=request.form["building_id"],
            capacity=capacity,
            occupied=request.form.get("occupied", 0),
            room_type=request.form.get("room_type", "四人间"),
            price=request.form.get("price", 1200.0),
        )
        db.session.add(r)
        db.session.commit()
        log_operation(f"添加房间「{r.room_number}」", "room", r.id)
        flash("添加成功", "success")
        return redirect(url_for("room_list"))
    buildings = Building.query.all()
    return render_template("room_form.html", room=None, buildings=buildings)


@app.route("/room/<int:rid>/edit", methods=["GET", "POST"])
@role_required("admin")
def room_edit(rid):
    r = Room.query.get_or_404(rid)
    if request.method == "POST":
        capacity = int(request.form.get("capacity", 4))
        if capacity < 1:
            flash("房间容量至少为1", "warning")
            return render_template("room_form.html", room=r, buildings=Building.query.all())
        r.room_number = request.form["room_number"]
        r.building_id = request.form["building_id"]
        r.capacity = capacity
        r.occupied = request.form.get("occupied", 0)
        r.room_type = request.form.get("room_type", "四人间")
        r.price = request.form.get("price", 1200.0)
        db.session.commit()
        log_operation(f"修改房间「{r.room_number}」", "room", r.id)
        flash("修改成功", "success")
        return redirect(url_for("room_list"))
    buildings = Building.query.all()
    return render_template("room_form.html", room=r, buildings=buildings)


@app.route("/room/<int:rid>/delete", methods=["POST"])
@role_required("admin")
def room_delete(rid):
    r = Room.query.get_or_404(rid)
    rn = r.room_number
    db.session.delete(r)
    db.session.commit()
    log_operation(f"删除房间「{rn}」", "room", rid)
    flash("删除成功", "success")
    return redirect(url_for("room_list"))


# ============================================================
# 学生管理
# ============================================================

@app.route("/students")
@role_required("admin")
def student_list():
    search = request.args.get("search", "").strip()
    page = request.args.get("page", 1, type=int)
    q = Student.query
    if search:
        q = q.filter(db.or_(Student.sno.contains(search), Student.name.contains(search), Student.major.contains(search), Student.class_name.contains(search), Student.phone.contains(search)))
    pagination = q.order_by(Student.sno).paginate(page=page, per_page=15)
    return render_template("students.html", students=pagination.items, pagination=pagination, search=search)


@app.route("/students/export")
@role_required("admin")
def student_export():
    search = request.args.get("search", "").strip()
    q = Student.query
    if search:
        q = q.filter(db.or_(Student.sno.contains(search), Student.name.contains(search), Student.major.contains(search), Student.class_name.contains(search), Student.phone.contains(search)))
    students = q.order_by(Student.sno).all()
    return export_excel(
        "学生列表",
        ["学号", "姓名", "性别", "出生日期", "手机", "专业", "班级"],
        [[s.sno, s.name, s.gender or "-", str(s.birth) if s.birth else "-", s.phone or "-", s.major or "-", s.class_name or "-"] for s in students],
    )


@app.route("/student/add", methods=["GET", "POST"])
@role_required("admin")
def student_add():
    if request.method == "POST":
        phone = request.form.get("phone", "").strip()
        if phone and not re.match(r"^1\d{10}$", phone):
            flash("手机号格式不正确（11位数字，以1开头）", "warning")
            return render_template("student_form.html", student=None)
        s = Student(
            sno=request.form["sno"],
            name=request.form["name"],
            gender=request.form.get("gender"),
            birth=request.form.get("birth") or None,
            phone=phone or None,
            major=request.form.get("major"),
            class_name=request.form.get("class_name"),
        )
        db.session.add(s)
        db.session.commit()
        log_operation(f"添加学生「{s.name}」({s.sno})", "student", s.id)
        flash("添加成功", "success")
        return redirect(url_for("student_list"))
    return render_template("student_form.html", student=None)


@app.route("/student/<int:sid>/edit", methods=["GET", "POST"])
@role_required("admin")
def student_edit(sid):
    s = Student.query.get_or_404(sid)
    if request.method == "POST":
        phone = request.form.get("phone", "").strip()
        if phone and not re.match(r"^1\d{10}$", phone):
            flash("手机号格式不正确（11位数字，以1开头）", "warning")
            return render_template("student_form.html", student=s)
        s.sno = request.form["sno"]
        s.name = request.form["name"]
        s.gender = request.form.get("gender")
        s.birth = request.form.get("birth") or None
        s.phone = phone or None
        s.major = request.form.get("major")
        s.class_name = request.form.get("class_name")
        db.session.commit()
        log_operation(f"修改学生「{s.name}」({s.sno})", "student", s.id)
        flash("修改成功", "success")
        return redirect(url_for("student_list"))
    return render_template("student_form.html", student=s)


@app.route("/student/<int:sid>/delete", methods=["POST"])
@role_required("admin")
def student_delete(sid):
    s = Student.query.get_or_404(sid)
    name, sno = s.name, s.sno
    db.session.delete(s)
    db.session.commit()
    log_operation(f"删除学生「{name}」({sno})", "student", sid)
    flash("删除成功", "success")
    return redirect(url_for("student_list"))


@app.route("/students/import", methods=["GET", "POST"])
@role_required("admin")
def student_import():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename:
            flash("请选择文件", "warning")
            return redirect(url_for("student_import"))

        ext = file.filename.rsplit(".", 1)[-1].lower()
        if ext not in ("xlsx", "xls"):
            flash("仅支持 .xlsx 或 .xls 格式", "danger")
            return redirect(url_for("student_import"))

        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
        except Exception as e:
            flash(f"无法解析 Excel 文件: {e}", "danger")
            return redirect(url_for("student_import"))

        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            flash("文件为空", "warning")
            return redirect(url_for("student_import"))

        headers = [str(h).strip() if h else "" for h in rows[0]]
        col_map = {}
        for i, h in enumerate(headers):
            if h in ("学号", "sno"):
                col_map["sno"] = i
            elif h in ("姓名", "name"):
                col_map["name"] = i
            elif h in ("性别", "gender"):
                col_map["gender"] = i
            elif h in ("出生日期", "birth"):
                col_map["birth"] = i
            elif h in ("手机", "phone"):
                col_map["phone"] = i
            elif h in ("专业", "major"):
                col_map["major"] = i
            elif h in ("班级", "class_name", "班级名称"):
                col_map["class_name"] = i

        if "sno" not in col_map or "name" not in col_map:
            flash("表头必须包含'学号'和'姓名'列", "danger")
            return redirect(url_for("student_import"))

        success = 0
        fail = 0
        errors = []
        seen_snos = set()

        for row_idx, row in enumerate(rows[1:], start=2):
            sno = str(row[col_map["sno"]]).strip() if row[col_map["sno"]] is not None else ""
            name = str(row[col_map["name"]]).strip() if row[col_map["name"]] is not None else ""

            if not sno or not name:
                fail += 1
                errors.append(f"第{row_idx}行: 学号或姓名为空")
                continue

            if sno in seen_snos:
                fail += 1
                errors.append(f"第{row_idx}行: 学号 {sno} 与文件中其他行重复")
                continue
            seen_snos.add(sno)

            if Student.query.filter_by(sno=sno).first():
                fail += 1
                errors.append(f"第{row_idx}行: 学号 {sno} 已存在")
                continue

            gender = str(row[col_map["gender"]]).strip() if "gender" in col_map and row[col_map["gender"]] is not None else None
            phone = str(row[col_map["phone"]]).strip() if "phone" in col_map and row[col_map["phone"]] is not None else None
            major = str(row[col_map["major"]]).strip() if "major" in col_map and row[col_map["major"]] is not None else None
            class_name = str(row[col_map["class_name"]]).strip() if "class_name" in col_map and row[col_map["class_name"]] is not None else None

            # Parse birth date
            birth = None
            if "birth" in col_map and row[col_map["birth"]] is not None:
                raw = row[col_map["birth"]]
                if isinstance(raw, datetime):
                    birth = raw.date()
                elif isinstance(raw, date):
                    birth = raw
                elif isinstance(raw, str):
                    raw = raw.strip()
                    if raw:
                        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
                            try:
                                birth = datetime.strptime(raw, fmt).date()
                                break
                            except ValueError:
                                pass

            try:
                s = Student(
                    sno=sno,
                    name=name,
                    gender=gender,
                    birth=birth,
                    phone=phone,
                    major=major,
                    class_name=class_name,
                )
                db.session.add(s)
                db.session.flush()
            except Exception as e:
                fail += 1
                errors.append(f"第{row_idx}行: 数据库错误 - {e}")
                db.session.rollback()
                continue

            success += 1

        db.session.commit()
        log_operation(f"批量导入学生（{success}成功 {fail}失败）", "student")
        results = {"success": success, "fail": fail, "errors": errors}
        return render_template("student_import.html", results=results)

    return render_template("student_import.html", results=None)


# ============================================================
# 入住管理
# ============================================================

@app.route("/accommodations")
@role_required("admin", "dorm_manager")
def accommodation_list():
    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip()
    page = request.args.get("page", 1, type=int)
    manager_bid = get_manager_building_id()
    q = Accommodation.query.options(
        db.joinedload(Accommodation.student),
        db.joinedload(Accommodation.room).joinedload(Room.building),
    )
    if manager_bid:
        q = q.join(Accommodation.room).filter(Room.building_id == manager_bid)
    if search:
        q = q.join(Accommodation.student).filter(Student.name.contains(search))
    if status_filter:
        q = q.filter(Accommodation.status == status_filter)
    pagination = q.order_by(Accommodation.check_in_date.desc()).paginate(page=page, per_page=15)
    return render_template("accommodations.html", accs=pagination.items, pagination=pagination, search=search, status_filter=status_filter)


@app.route("/accommodations/export")
@role_required("admin", "dorm_manager")
def accommodation_export():
    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip()
    manager_bid = get_manager_building_id()
    q = Accommodation.query.options(
        db.joinedload(Accommodation.student),
        db.joinedload(Accommodation.room).joinedload(Room.building),
    )
    if manager_bid:
        q = q.join(Accommodation.room).filter(Room.building_id == manager_bid)
    if search:
        q = q.join(Accommodation.student).filter(Student.name.contains(search))
    if status_filter:
        q = q.filter(Accommodation.status == status_filter)
    accs = q.order_by(Accommodation.check_in_date.desc()).all()
    return export_excel(
        "入住记录",
        ["学生学号", "学生姓名", "宿舍楼", "房间号", "入住日期", "退宿日期", "状态"],
        [[a.student.sno, a.student.name, a.room.building.name, a.room.room_number, str(a.check_in_date), str(a.check_out_date) if a.check_out_date else "-", a.status] for a in accs],
    )


@app.route("/accommodation/add", methods=["GET", "POST"])
@role_required("admin", "dorm_manager")
def accommodation_add():
    if request.method == "POST":
        student_id = request.form["student_id"]
        room_id = request.form["room_id"]
        check_in_date = request.form["check_in_date"]
        try:
            db.session.execute(
                text("CALL sp_checkin(:sid, :rid, :cdate)"),
                {"sid": student_id, "rid": room_id, "cdate": check_in_date},
            )
            db.session.commit()
            log_operation(f"为学生登记入住", "accommodation")
            flash("入住登记成功", "success")
        except Exception as e:
            db.session.rollback()
            # 提取 MySQL SIGNAL 消息
            msg = str(e)
            if "该学生已入住" in msg:
                flash("该学生已入住，请先退宿", "danger")
            elif "该房间已满" in msg:
                flash("该房间已满", "danger")
            else:
                flash(f"操作失败: {msg}", "danger")
            return redirect(url_for("accommodation_add"))
        return redirect(url_for("accommodation_list"))
    manager_bid = get_manager_building_id()
    students = Student.query.order_by(Student.sno).all()
    rooms = Room.query
    if manager_bid:
        rooms = rooms.filter_by(building_id=manager_bid)
    rooms = rooms.order_by(Room.building_id, Room.room_number).all()
    return render_template("accommodation_form.html", acc=None, students=students, rooms=rooms)


@app.route("/accommodation/<int:aid>/checkout", methods=["POST"])
@role_required("admin", "dorm_manager")
def accommodation_checkout(aid):
    acc = Accommodation.query.get_or_404(aid)
    if acc.status == "已退宿":
        flash("该记录已退宿", "warning")
        return redirect(url_for("accommodation_list"))
    acc.status = "已退宿"
    acc.check_out_date = date.today()
    # trg_accommodation_checkout 触发器自动更新 room.occupied
    db.session.commit()
    log_operation(f"办理退宿", "accommodation", aid)
    flash("退宿成功", "success")
    return redirect(url_for("accommodation_list"))


# ============================================================
# 退宿审核
# ============================================================

@app.route("/my/accommodation/checkout-request", methods=["POST"])
@login_required
def checkout_request():
    """学生提交退宿申请"""
    student = Student.query.filter_by(user_id=session["user_id"]).first()
    if not student:
        flash("未关联学生信息", "warning")
        return redirect(url_for("index"))
    existing = CheckoutRequest.query.filter_by(
        student_id=student.id, status="pending"
    ).first()
    if existing:
        flash("已有退宿申请正在审核中", "warning")
        return redirect(url_for("my_accommodation"))
    acc = Accommodation.query.filter_by(
        student_id=student.id, status="入住"
    ).first()
    if not acc:
        flash("当前无在住记录", "warning")
        return redirect(url_for("my_accommodation"))
    req = CheckoutRequest(student_id=student.id, accommodation_id=acc.id)
    db.session.add(req)
    db.session.commit()
    flash("退宿申请已提交，请等待宿管审核", "success")
    return redirect(url_for("my_accommodation"))


@app.route("/checkout-requests")
@role_required("admin", "dorm_manager")
def checkout_request_list():
    manager_bid = get_manager_building_id()
    page = request.args.get("page", 1, type=int)
    q = CheckoutRequest.query.options(
        db.joinedload(CheckoutRequest.student),
        db.joinedload(CheckoutRequest.accommodation)
        .joinedload(Accommodation.room)
        .joinedload(Room.building),
        db.joinedload(CheckoutRequest.reviewer),
    )
    if manager_bid:
        q = q.join(CheckoutRequest.accommodation)\
             .join(Accommodation.room)\
             .filter(Room.building_id == manager_bid)
    pagination = q.order_by(
        db.case((CheckoutRequest.status == "pending", 0), else_=1),
        CheckoutRequest.request_date.desc()
    ).paginate(page=page, per_page=15)
    return render_template("checkout_requests.html",
                           requests=pagination.items,
                           pagination=pagination)


@app.route("/checkout-requests/export")
@role_required("admin", "dorm_manager")
def checkout_request_export():
    manager_bid = get_manager_building_id()
    q = CheckoutRequest.query.options(
        db.joinedload(CheckoutRequest.student),
        db.joinedload(CheckoutRequest.accommodation)
        .joinedload(Accommodation.room)
        .joinedload(Room.building),
        db.joinedload(CheckoutRequest.reviewer),
    )
    if manager_bid:
        q = q.join(CheckoutRequest.accommodation)\
             .join(Accommodation.room)\
             .filter(Room.building_id == manager_bid)
    reqs = q.order_by(CheckoutRequest.request_date.desc()).all()
    return export_excel(
        "退宿申请列表",
        ["学生学号", "学生姓名", "宿舍楼", "房间号", "入住日期", "申请日期", "状态", "审核人", "审核日期"],
        [[r.student.sno, r.student.name, r.accommodation.room.building.name,
          r.accommodation.room.room_number, str(r.accommodation.check_in_date),
          r.request_date.strftime("%Y-%m-%d %H:%M") if r.request_date else "",
          {"pending": "待审核", "approved": "已批准", "rejected": "已拒绝"}.get(r.status, r.status),
          r.reviewer.username if r.reviewer else "-",
          r.review_date.strftime("%Y-%m-%d %H:%M") if r.review_date else "-"]
         for r in reqs],
    )


@app.route("/checkout-request/<int:rid>/approve", methods=["POST"])
@role_required("admin", "dorm_manager")
def checkout_request_approve(rid):
    req = CheckoutRequest.query.get_or_404(rid)
    if req.status != "pending":
        flash("该申请已处理", "warning")
        return redirect(url_for("checkout_request_list"))
    acc = Accommodation.query.get(req.accommodation_id)
    manager_bid = get_manager_building_id()
    if manager_bid and acc.room.building_id != manager_bid:
        flash("无权处理该申请", "danger")
        return redirect(url_for("checkout_request_list"))
    acc.status = "已退宿"
    acc.check_out_date = date.today()
    req.status = "approved"
    req.reviewed_by = session["user_id"]
    req.review_date = datetime.now()
    db.session.commit()
    log_operation(f"批准退宿申请 #{rid}", "checkout_request", rid)
    flash("退宿申请已批准", "success")
    return redirect(url_for("checkout_request_list"))


@app.route("/checkout-request/<int:rid>/reject", methods=["POST"])
@role_required("admin", "dorm_manager")
def checkout_request_reject(rid):
    req = CheckoutRequest.query.get_or_404(rid)
    if req.status != "pending":
        flash("该申请已处理", "warning")
        return redirect(url_for("checkout_request_list"))
    acc = Accommodation.query.get(req.accommodation_id)
    manager_bid = get_manager_building_id()
    if manager_bid and acc.room.building_id != manager_bid:
        flash("无权处理该申请", "danger")
        return redirect(url_for("checkout_request_list"))
    req.status = "rejected"
    req.reviewed_by = session["user_id"]
    req.review_date = datetime.now()
    db.session.commit()
    log_operation(f"拒绝退宿申请 #{rid}", "checkout_request", rid)
    flash("退宿申请已拒绝", "success")
    return redirect(url_for("checkout_request_list"))


# ============================================================
# 宿舍调换
# ============================================================

@app.route("/my/accommodation/transfer-request", methods=["POST"])
@login_required
def transfer_request():
    """学生提交调换申请"""
    student = Student.query.filter_by(user_id=session["user_id"]).first()
    if not student:
        flash("未关联学生信息", "warning")
        return redirect(url_for("index"))
    # 检查是否有待审核的调换申请
    existing = TransferRequest.query.filter_by(
        student_id=student.id, status="pending"
    ).first()
    if existing:
        flash("已有调换申请正在审核中", "warning")
        return redirect(url_for("my_accommodation"))
    # 检查是否有在住记录
    acc = Accommodation.query.filter_by(
        student_id=student.id, status="入住"
    ).first()
    if not acc:
        flash("当前无在住记录", "warning")
        return redirect(url_for("my_accommodation"))
    to_room_id = request.form.get("to_room_id", type=int)
    if not to_room_id:
        flash("请选择目标房间", "warning")
        return redirect(url_for("my_accommodation"))
    # 校验目标房间有空位
    target_room = Room.query.get_or_404(to_room_id)
    if target_room.occupied >= target_room.capacity:
        flash("目标房间已满", "warning")
        return redirect(url_for("my_accommodation"))
    if target_room.id == acc.room_id:
        flash("不能调换到当前房间", "warning")
        return redirect(url_for("my_accommodation"))
    req = TransferRequest(
        student_id=student.id,
        from_accommodation_id=acc.id,
        to_room_id=to_room_id,
    )
    db.session.add(req)
    db.session.commit()
    flash("调换申请已提交，请等待宿管审核", "success")
    return redirect(url_for("my_accommodation"))


@app.route("/transfer-requests")
@role_required("admin", "dorm_manager")
def transfer_request_list():
    manager_bid = get_manager_building_id()
    page = request.args.get("page", 1, type=int)
    q = TransferRequest.query.options(
        db.joinedload(TransferRequest.student),
        db.joinedload(TransferRequest.from_accommodation)
        .joinedload(Accommodation.room)
        .joinedload(Room.building),
        db.joinedload(TransferRequest.to_room).joinedload(Room.building),
        db.joinedload(TransferRequest.reviewer),
    )
    if manager_bid:
        q = q.join(TransferRequest.from_accommodation)\
             .join(Accommodation.room)\
             .filter(Room.building_id == manager_bid)
    pagination = q.order_by(
        db.case((TransferRequest.status == "pending", 0), else_=1),
        TransferRequest.request_date.desc()
    ).paginate(page=page, per_page=15)
    return render_template("transfer_requests.html",
                           requests=pagination.items,
                           pagination=pagination)


@app.route("/transfer-requests/export")
@role_required("admin", "dorm_manager")
def transfer_request_export():
    manager_bid = get_manager_building_id()
    q = TransferRequest.query.options(
        db.joinedload(TransferRequest.student),
        db.joinedload(TransferRequest.from_accommodation)
        .joinedload(Accommodation.room)
        .joinedload(Room.building),
        db.joinedload(TransferRequest.to_room).joinedload(Room.building),
        db.joinedload(TransferRequest.reviewer),
    )
    if manager_bid:
        q = q.join(TransferRequest.from_accommodation)\
             .join(Accommodation.room)\
             .filter(Room.building_id == manager_bid)
    reqs = q.order_by(TransferRequest.request_date.desc()).all()
    return export_excel(
        "调换申请列表",
        ["学生学号", "学生姓名", "原宿舍楼", "原房间号", "目标宿舍楼", "目标房间号", "申请日期", "状态", "审核人", "审核日期"],
        [[r.student.sno, r.student.name,
          r.from_accommodation.room.building.name, r.from_accommodation.room.room_number,
          r.to_room.building.name, r.to_room.room_number,
          r.request_date.strftime("%Y-%m-%d %H:%M") if r.request_date else "",
          {"pending": "待审核", "approved": "已批准", "rejected": "已拒绝"}.get(r.status, r.status),
          r.reviewer.username if r.reviewer else "-",
          r.review_date.strftime("%Y-%m-%d %H:%M") if r.review_date else "-"]
         for r in reqs],
    )


@app.route("/transfer-request/<int:rid>/approve", methods=["POST"])
@role_required("admin", "dorm_manager")
def transfer_request_approve(rid):
    req = TransferRequest.query.get_or_404(rid)
    if req.status != "pending":
        flash("该申请已处理", "warning")
        return redirect(url_for("transfer_request_list"))
    # 校验管辖范围
    acc = Accommodation.query.get(req.from_accommodation_id)
    manager_bid = get_manager_building_id()
    if manager_bid and acc.room.building_id != manager_bid:
        flash("无权处理该申请", "danger")
        return redirect(url_for("transfer_request_list"))
    try:
        db.session.execute(
            text("CALL sp_room_transfer(:aid, :rid)"),
            {"aid": req.from_accommodation_id, "rid": req.to_room_id},
        )
        req.status = "approved"
        req.reviewed_by = session["user_id"]
        req.review_date = datetime.now()
        db.session.commit()
        log_operation(f"批准调换申请 #{rid}", "transfer_request", rid)
        flash("调换申请已批准，房间已迁移", "success")
    except Exception as e:
        db.session.rollback()
        msg = str(e)
        if "不能调换到同一房间" in msg:
            flash("不能调换到同一房间", "danger")
        elif "目标房间已满" in msg:
            flash("目标房间已满", "danger")
        else:
            flash(f"操作失败: {msg}", "danger")
    return redirect(url_for("transfer_request_list"))


@app.route("/transfer-request/<int:rid>/reject", methods=["POST"])
@role_required("admin", "dorm_manager")
def transfer_request_reject(rid):
    req = TransferRequest.query.get_or_404(rid)
    if req.status != "pending":
        flash("该申请已处理", "warning")
        return redirect(url_for("transfer_request_list"))
    acc = Accommodation.query.get(req.from_accommodation_id)
    manager_bid = get_manager_building_id()
    if manager_bid and acc.room.building_id != manager_bid:
        flash("无权处理该申请", "danger")
        return redirect(url_for("transfer_request_list"))
    req.status = "rejected"
    req.reviewed_by = session["user_id"]
    req.review_date = datetime.now()
    db.session.commit()
    log_operation(f"拒绝调换申请 #{rid}", "transfer_request", rid)
    flash("调换申请已拒绝", "success")
    return redirect(url_for("transfer_request_list"))


# ============================================================
# 报修管理
# ============================================================

@app.route("/repairs")
@role_required("admin", "dorm_manager")
def repair_list():
    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip()
    page = request.args.get("page", 1, type=int)
    manager_bid = get_manager_building_id()
    q = Repair.query.options(
        db.joinedload(Repair.student),
        db.joinedload(Repair.room).joinedload(Room.building),
    )
    if manager_bid:
        q = q.join(Repair.room).filter(Room.building_id == manager_bid)
    if search:
        q = q.join(Repair.student).filter(db.or_(Repair.description.contains(search), Student.name.contains(search)))
    if status_filter:
        q = q.filter(Repair.status == status_filter)
    pagination = q.order_by(Repair.report_date.desc()).paginate(page=page, per_page=15)
    return render_template("repairs.html", repairs=pagination.items, pagination=pagination, search=search, status_filter=status_filter)


@app.route("/repairs/export")
@role_required("admin", "dorm_manager")
def repair_export():
    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip()
    manager_bid = get_manager_building_id()
    q = Repair.query.options(
        db.joinedload(Repair.student),
        db.joinedload(Repair.room).joinedload(Room.building),
    )
    if manager_bid:
        q = q.join(Repair.room).filter(Room.building_id == manager_bid)
    if search:
        q = q.join(Repair.student).filter(db.or_(Repair.description.contains(search), Student.name.contains(search)))
    if status_filter:
        q = q.filter(Repair.status == status_filter)
    repairs = q.order_by(Repair.report_date.desc()).all()
    return export_excel(
        "报修列表",
        ["学生学号", "学生姓名", "宿舍楼", "房间号", "描述", "报修日期", "完成日期", "状态"],
        [[r.student.sno, r.student.name, r.room.building.name, r.room.room_number, r.description, str(r.report_date), str(r.fix_date) if r.fix_date else "-", r.status] for r in repairs],
    )


@app.route("/repair/add", methods=["GET", "POST"])
@login_required
def repair_add():
    if request.method == "POST":
        # 处理上传文件
        filename = None
        file = request.files.get("image")
        if file and file.filename and allowed_file(file.filename):
            ext = file.filename.rsplit(".", 1)[1].lower()
            filename = f"{uuid.uuid4().hex}.{ext}"
            file.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))

        r = Repair(
            student_id=request.form["student_id"],
            room_id=request.form["room_id"],
            description=request.form["description"],
            image=filename,
            status="待处理",
            report_date=request.form.get("report_date", date.today()),
        )
        db.session.add(r)
        db.session.commit()
        flash("报修提交成功", "success")
        return redirect(url_for("repair_list"))
    students = Student.query.order_by(Student.sno).all()
    rooms = Room.query
    if get_manager_building_id():
        rooms = rooms.filter_by(building_id=get_manager_building_id())
    rooms = rooms.order_by(Room.building_id, Room.room_number).all()
    return render_template("repair_form.html", repair=None, students=students, rooms=rooms)


@app.route("/repair/<int:rid>/status", methods=["POST"])
@role_required("admin", "dorm_manager")
def repair_status(rid):
    r = Repair.query.get_or_404(rid)
    new_status = request.form["status"]
    r.status = new_status
    # trg_repair_complete 触发器自动设置 fix_date
    # 可选重新上传照片
    file = request.files.get("image")
    if file and file.filename and allowed_file(file.filename):
        ext = file.filename.rsplit(".", 1)[1].lower()
        filename = f"{uuid.uuid4().hex}.{ext}"
        file.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))
        r.image = filename
    db.session.commit()
    log_operation(f"更新报修状态为「{new_status}」", "repair", rid)
    flash("状态更新成功", "success")
    return redirect(url_for("repair_list"))


@app.route("/uploads/<filename>")
def uploaded_file(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


# ============================================================
# 访客管理
# ============================================================

@app.route("/visitors")
@role_required("admin", "dorm_manager")
def visitor_list():
    search = request.args.get("search", "").strip()
    page = request.args.get("page", 1, type=int)
    manager_bid = get_manager_building_id()
    q = Visitor.query.options(db.joinedload(Visitor.student))
    if manager_bid:
        q = q.join(Visitor.student).join(Accommodation, db.and_(Accommodation.student_id == Student.id, Accommodation.status == "入住")).join(Room).filter(Room.building_id == manager_bid)
    if search:
        q = q.join(Visitor.student).filter(db.or_(Visitor.visitor_name.contains(search), Visitor.id_card.contains(search), Student.name.contains(search)))
    pagination = q.order_by(Visitor.visit_date.desc()).paginate(page=page, per_page=15)
    return render_template("visitors.html", visitors=pagination.items, pagination=pagination, search=search)


@app.route("/visitors/export")
@role_required("admin", "dorm_manager")
def visitor_export():
    search = request.args.get("search", "").strip()
    manager_bid = get_manager_building_id()
    q = Visitor.query.options(db.joinedload(Visitor.student))
    if manager_bid:
        q = q.join(Visitor.student).join(Accommodation, db.and_(Accommodation.student_id == Student.id, Accommodation.status == "入住")).join(Room).filter(Room.building_id == manager_bid)
    if search:
        q = q.join(Visitor.student).filter(db.or_(Visitor.visitor_name.contains(search), Visitor.id_card.contains(search), Student.name.contains(search)))
    visitors = q.order_by(Visitor.visit_date.desc()).all()
    return export_excel(
        "访客列表",
        ["受访学生学号", "受访学生姓名", "访客姓名", "身份证号", "事由", "来访时间", "离开时间"],
        [[v.student.sno, v.student.name, v.visitor_name, v.id_card, v.reason or "-", str(v.visit_date), str(v.leave_date) if v.leave_date else "-"] for v in visitors],
    )


@app.route("/visitor/add", methods=["GET", "POST"])
@login_required
def visitor_add():
    if request.method == "POST":
        id_card = request.form.get("id_card", "").strip()
        if not re.match(r"^\d{17}[\dXx]$", id_card):
            flash("身份证号格式不正确（18位）", "warning")
            return render_template("visitor_form.html", visitor=None, students=Student.query.order_by(Student.sno).all())
        v = Visitor(
            student_id=request.form["student_id"],
            visitor_name=request.form["visitor_name"],
            id_card=id_card,
            reason=request.form.get("reason"),
            visit_date=request.form["visit_date"],
            leave_date=request.form.get("leave_date") or None,
        )
        db.session.add(v)
        db.session.commit()
        flash("访客登记成功", "success")
        return redirect(url_for("visitor_list"))
    students = Student.query
    if get_manager_building_id():
        students = students.join(Accommodation, db.and_(Accommodation.student_id == Student.id, Accommodation.status == "入住")).join(Room).filter(Room.building_id == get_manager_building_id())
    students = students.order_by(Student.sno).all()
    return render_template("visitor_form.html", visitor=None, students=students)


# ============================================================
# 费用管理
# ============================================================

@app.route("/fees")
@role_required("admin", "dorm_manager")
def fee_list():
    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip()
    page = request.args.get("page", 1, type=int)
    manager_bid = get_manager_building_id()
    q = Fee.query.options(
        db.joinedload(Fee.student),
        db.joinedload(Fee.room).joinedload(Room.building),
    )
    if manager_bid:
        q = q.join(Fee.room).filter(Room.building_id == manager_bid)
    if search:
        q = q.join(Fee.student).filter(db.or_(Fee.fee_type.contains(search), Student.name.contains(search)))
    if status_filter:
        q = q.filter(Fee.status == status_filter)
    pagination = q.order_by(Fee.due_date.desc()).paginate(page=page, per_page=15)
    buildings = Building.query.all() if not manager_bid else Building.query.filter_by(id=manager_bid).all()
    return render_template("fees.html", fees=pagination.items, pagination=pagination, search=search, status_filter=status_filter, buildings=buildings)


@app.route("/fees/export")
@role_required("admin", "dorm_manager")
def fee_export():
    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip()
    manager_bid = get_manager_building_id()
    q = Fee.query.options(
        db.joinedload(Fee.student),
        db.joinedload(Fee.room).joinedload(Room.building),
    )
    if manager_bid:
        q = q.join(Fee.room).filter(Room.building_id == manager_bid)
    if search:
        q = q.join(Fee.student).filter(db.or_(Fee.fee_type.contains(search), Student.name.contains(search)))
    if status_filter:
        q = q.filter(Fee.status == status_filter)
    fees = q.order_by(Fee.due_date.desc()).all()
    return export_excel(
        "费用列表",
        ["学生学号", "学生姓名", "宿舍楼", "房间号", "费用类型", "金额", "截止日期", "缴费日期", "状态"],
        [[f.student.sno, f.student.name, f.room.building.name, f.room.room_number, f.fee_type, f.amount, str(f.due_date) if f.due_date else "-", str(f.pay_date) if f.pay_date else "-", f.status] for f in fees],
    )


@app.route("/fee/add", methods=["GET", "POST"])
@role_required("admin", "dorm_manager")
def fee_add():
    if request.method == "POST":
        try:
            amount = float(request.form["amount"])
            if amount <= 0:
                raise ValueError
        except ValueError:
            flash("金额必须为正数", "warning")
            return render_template("fee_form.html", fee=None, students=Student.query.order_by(Student.sno).all(), rooms=Room.query.order_by(Room.building_id, Room.room_number).all())
        f = Fee(
            student_id=request.form["student_id"],
            room_id=request.form["room_id"],
            fee_type=request.form["fee_type"],
            amount=amount,
            status="未缴",
            due_date=request.form.get("due_date") or None,
        )
        db.session.add(f)
        db.session.commit()
        log_operation(f"添加费用「{f.fee_type}」¥{f.amount}", "fee", f.id)
        flash("费用添加成功", "success")
        return redirect(url_for("fee_list"))
    students = Student.query
    rooms = Room.query
    if get_manager_building_id():
        rooms = rooms.filter_by(building_id=get_manager_building_id())
    rooms = rooms.order_by(Room.building_id, Room.room_number).all()
    students = students.order_by(Student.sno).all()
    return render_template("fee_form.html", fee=None, students=students, rooms=rooms)


@app.route("/fee/<int:fid>/pay", methods=["POST"])
@role_required("admin", "dorm_manager")
def fee_pay(fid):
    f = Fee.query.get_or_404(fid)
    if f.status == "已缴":
        flash("该费用已缴纳", "warning")
        return redirect(url_for("fee_list"))
    f.status = "已缴"
    f.pay_date = date.today()
    db.session.commit()
    log_operation(f"标记费用已缴「{f.fee_type}」¥{f.amount}", "fee", fid)
    flash("缴费成功", "success")
    return redirect(url_for("fee_list"))


@app.route("/fees/batch", methods=["POST"])
@role_required("admin", "dorm_manager")
def fee_batch():
    building_id = request.form["building_id"]
    # dorm_manager 只能为自己管辖的楼生成费用
    manager_bid = get_manager_building_id()
    if manager_bid and int(building_id) != manager_bid:
        flash("只能为自己管辖的宿舍楼批量生成费用", "danger")
        return redirect(url_for("fee_list"))
    fee_type = request.form["fee_type"]
    amount = request.form["amount"]
    due_date = request.form.get("due_date") or None
    try:
        result = db.session.execute(
            text("CALL sp_generate_fees(:bid, :ftype, :amt, :ddate)"),
            {"bid": building_id, "ftype": fee_type, "amt": amount, "ddate": due_date},
        )
        count = result.fetchone()[0]
        db.session.commit()
        log_operation(f"批量生成费用「{fee_type}」¥{amount}（{count}人）", "fee")
        flash(f"已为 {count} 名学生生成{fee_type}（{amount}元）", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"操作失败: {e}", "danger")
    return redirect(url_for("fee_list"))


# ============================================================
# 用户管理（管理员专属）
# ============================================================

@app.route("/users")
@role_required("admin")
def user_list():
    search = request.args.get("search", "").strip()
    page = request.args.get("page", 1, type=int)
    q = User.query
    if search:
        q = q.filter(User.username.contains(search))
    pagination = q.order_by(User.id).paginate(page=page, per_page=15)
    return render_template("users.html", users=pagination.items, pagination=pagination, search=search)


@app.route("/users/export")
@role_required("admin")
def user_export():
    search = request.args.get("search", "").strip()
    q = User.query.options(db.joinedload(User.building))
    if search:
        q = q.filter(User.username.contains(search))
    users = q.order_by(User.id).all()
    return export_excel(
        "用户列表",
        ["ID", "用户名", "角色", "管辖宿舍楼"],
        [[u.id, u.username, u.role, u.building.name if u.building else "-"] for u in users],
    )


@app.route("/user/add", methods=["GET", "POST"])
@role_required("admin")
def user_add():
    if request.method == "POST":
        password = request.form.get("password", "").strip()
        if len(password) < 4:
            flash("密码至少4位", "warning")
            return render_template("user_form.html", user=None, buildings=Building.query.all())
        building_id = request.form.get("building_id")
        u = User(
            username=request.form["username"],
            password=generate_password_hash(password),
            role=request.form["role"],
            building_id=int(building_id) if building_id and request.form["role"] == "dorm_manager" else None,
        )
        db.session.add(u)
        db.session.commit()
        log_operation(f"添加用户「{u.username}」({u.role})", "user", u.id)
        flash("用户添加成功", "success")
        return redirect(url_for("user_list"))
    buildings = Building.query.all()
    return render_template("user_form.html", user=None, buildings=buildings)


@app.route("/user/<int:uid>/edit", methods=["GET", "POST"])
@role_required("admin")
def user_edit(uid):
    u = User.query.get_or_404(uid)
    if request.method == "POST":
        u.username = request.form["username"]
        new_pw = request.form.get("password", "").strip()
        if new_pw:
            if len(new_pw) < 4:
                flash("密码至少4位", "warning")
                return render_template("user_form.html", user=u, buildings=Building.query.all())
            u.password = generate_password_hash(new_pw)
        u.role = request.form["role"]
        building_id = request.form.get("building_id")
        u.building_id = int(building_id) if building_id and request.form["role"] == "dorm_manager" else None
        db.session.commit()
        log_operation(f"修改用户「{u.username}」({u.role})", "user", u.id)
        flash("用户修改成功", "success")
        return redirect(url_for("user_list"))
    buildings = Building.query.all()
    return render_template("user_form.html", user=u, buildings=buildings)


@app.route("/user/<int:uid>/delete", methods=["POST"])
@role_required("admin")
def user_delete(uid):
    if uid == session["user_id"]:
        flash("不能删除自己", "danger")
        return redirect(url_for("user_list"))
    u = User.query.get_or_404(uid)
    uname, urole = u.username, u.role
    # 解除学生关联
    if u.student_profile:
        u.student_profile.user_id = None
    # 清除审核人引用
    CheckoutRequest.query.filter_by(reviewed_by=uid).update({"reviewed_by": None})
    TransferRequest.query.filter_by(reviewed_by=uid).update({"reviewed_by": None})
    # 清除该用户的公告和操作日志
    Announcement.query.filter_by(user_id=uid).delete()
    OperationLog.query.filter_by(user_id=uid).delete()
    db.session.flush()
    db.session.delete(u)
    db.session.commit()
    log_operation(f"删除用户「{uname}」({urole})", "user", uid)
    flash("用户删除成功", "success")
    return redirect(url_for("user_list"))


# ============================================================
# 公告管理
# ============================================================

@app.route("/announcements")
@login_required
def announcement_list():
    search = request.args.get("search", "").strip()
    page = request.args.get("page", 1, type=int)
    q = Announcement.query.options(db.joinedload(Announcement.user))
    if search:
        q = q.filter(db.or_(Announcement.title.contains(search), Announcement.content.contains(search)))
    pagination = q.order_by(Announcement.created_at.desc()).paginate(page=page, per_page=15)
    return render_template("announcements.html", announcements=pagination.items, pagination=pagination, search=search)


@app.route("/announcements/export")
@login_required
def announcement_export():
    search = request.args.get("search", "").strip()
    q = Announcement.query.options(db.joinedload(Announcement.user))
    if search:
        q = q.filter(db.or_(Announcement.title.contains(search), Announcement.content.contains(search)))
    announcements = q.order_by(Announcement.created_at.desc()).all()
    return export_excel(
        "公告列表",
        ["标题", "内容", "发布人", "发布时间"],
        [[a.title, a.content, a.user.username, a.created_at.strftime("%Y-%m-%d %H:%M")] for a in announcements],
    )


@app.route("/announcement/add", methods=["GET", "POST"])
@role_required("admin", "dorm_manager")
def announcement_add():
    if request.method == "POST":
        a = Announcement(
            title=request.form["title"],
            content=request.form["content"],
            user_id=session["user_id"],
        )
        db.session.add(a)
        db.session.commit()
        log_operation(f"发布公告「{a.title}」", "announcement", a.id)
        flash("公告发布成功", "success")
        return redirect(url_for("announcement_list"))
    return render_template("announcement_form.html", announcement=None)


@app.route("/announcement/<int:aid>/edit", methods=["GET", "POST"])
@role_required("admin", "dorm_manager")
def announcement_edit(aid):
    a = Announcement.query.get_or_404(aid)
    if request.method == "POST":
        a.title = request.form["title"]
        a.content = request.form["content"]
        db.session.commit()
        log_operation(f"修改公告「{a.title}」", "announcement", a.id)
        flash("公告修改成功", "success")
        return redirect(url_for("announcement_list"))
    return render_template("announcement_form.html", announcement=a)


@app.route("/announcement/<int:aid>/delete", methods=["POST"])
@role_required("admin", "dorm_manager")
def announcement_delete(aid):
    a = Announcement.query.get_or_404(aid)
    title = a.title
    db.session.delete(a)
    db.session.commit()
    log_operation(f"删除公告「{title}」", "announcement", aid)
    flash("公告删除成功", "success")
    return redirect(url_for("announcement_list"))


# ============================================================
# 操作日志
# ============================================================

@app.route("/operation-logs")
@role_required("admin")
def operation_log_list():
    search = request.args.get("search", "").strip()
    page = request.args.get("page", 1, type=int)
    q = OperationLog.query.options(db.joinedload(OperationLog.user))
    if search:
        q = q.join(OperationLog.user).filter(
            db.or_(User.username.contains(search), OperationLog.action.contains(search))
        )
    pagination = q.order_by(OperationLog.created_at.desc()).paginate(page=page, per_page=20)
    return render_template("operation_logs.html", logs=pagination.items, pagination=pagination, search=search)


@app.route("/operation-logs/export")
@role_required("admin")
def operation_log_export():
    search = request.args.get("search", "").strip()
    q = OperationLog.query.options(db.joinedload(OperationLog.user))
    if search:
        q = q.join(OperationLog.user).filter(
            db.or_(User.username.contains(search), OperationLog.action.contains(search))
        )
    logs = q.order_by(OperationLog.created_at.desc()).all()
    return export_excel(
        "操作日志",
        ["时间", "操作人", "角色", "操作内容", "IP地址"],
        [[log.created_at.strftime("%Y-%m-%d %H:%M:%S"), log.user.username, log.user.role, log.action, log.ip_address or "-"] for log in logs],
    )


# ============================================================
# 学生端
# ============================================================

@app.route("/my/accommodation")
@login_required
def my_accommodation():
    student = Student.query.filter_by(user_id=session["user_id"]).first()
    if not student:
        flash("未关联学生信息", "warning")
        return redirect(url_for("index"))
    acc = Accommodation.query.options(
        db.joinedload(Accommodation.room).joinedload(Room.building)
    ).filter_by(student_id=student.id, status="入住").first()
    roommates = []
    available_rooms = []
    if acc:
        roommates = Accommodation.query.options(
            db.joinedload(Accommodation.student)
        ).filter(
            Accommodation.room_id == acc.room_id,
            Accommodation.status == "入住",
            Accommodation.student_id != student.id,
        ).all()
        building_id = acc.room.building_id
        available_rooms = Room.query.filter(
            Room.building_id == building_id,
            Room.id != acc.room_id,
            Room.occupied < Room.capacity,
        ).order_by(Room.room_number).all()
    return render_template("my_accommodation.html", acc=acc, roommates=roommates, available_rooms=available_rooms)


@app.route("/my/repairs", methods=["GET", "POST"])
@login_required
def my_repairs():
    student = Student.query.filter_by(user_id=session["user_id"]).first()
    if not student:
        flash("未关联学生信息", "warning")
        return redirect(url_for("index"))
    if request.method == "POST":
        # 提交新报修
        filename = None
        file = request.files.get("image")
        if file and file.filename and allowed_file(file.filename):
            ext = file.filename.rsplit(".", 1)[1].lower()
            filename = f"{uuid.uuid4().hex}.{ext}"
            file.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))
        r = Repair(
            student_id=student.id,
            room_id=request.form["room_id"],
            description=request.form["description"],
            image=filename,
            status="待处理",
            report_date=date.today(),
        )
        db.session.add(r)
        db.session.commit()
        flash("报修提交成功", "success")
        return redirect(url_for("my_repairs"))
    repairs = Repair.query.filter_by(student_id=student.id).order_by(Repair.report_date.desc()).all()
    # 获取学生的入住房间用于报修表单
    my_rooms = Room.query.join(Accommodation).filter(
        Accommodation.student_id == student.id,
        Accommodation.status == "入住",
    ).all()
    return render_template("my_repairs.html", repairs=repairs, my_rooms=my_rooms)


@app.route("/my/repairs/<int:rid>/cancel", methods=["POST"])
@login_required
def my_repair_cancel(rid):
    r = Repair.query.get_or_404(rid)
    student = Student.query.filter_by(user_id=session["user_id"]).first()
    if not student or r.student_id != student.id:
        flash("无权操作", "danger")
        return redirect(url_for("my_repairs"))
    if r.status != "待处理":
        flash("仅可撤销待处理状态的报修", "warning")
        return redirect(url_for("my_repairs"))
    r.status = "已撤销"
    db.session.commit()
    flash("报修已撤销", "success")
    return redirect(url_for("my_repairs"))


@app.route("/my/fees")
@login_required
def my_fees():
    student = Student.query.filter_by(user_id=session["user_id"]).first()
    if not student:
        flash("未关联学生信息", "warning")
        return redirect(url_for("index"))
    fees = Fee.query.filter_by(student_id=student.id).order_by(Fee.due_date.desc()).all()
    return render_template("my_fees.html", fees=fees)


@app.route("/my/visitors", methods=["GET", "POST"])
@login_required
def my_visitors():
    student = Student.query.filter_by(user_id=session["user_id"]).first()
    if not student:
        flash("未关联学生信息", "warning")
        return redirect(url_for("index"))
    if request.method == "POST":
        id_card = request.form.get("id_card", "").strip()
        if not re.match(r"^\d{17}[\dXx]$", id_card):
            flash("身份证号格式不正确（18位）", "warning")
            return redirect(url_for("my_visitors"))
        v = Visitor(
            student_id=student.id,
            visitor_name=request.form["visitor_name"],
            id_card=id_card,
            reason=request.form.get("reason"),
            visit_date=request.form["visit_date"],
            leave_date=request.form.get("leave_date") or None,
        )
        db.session.add(v)
        db.session.commit()
        flash("访客登记成功", "success")
        return redirect(url_for("my_visitors"))
    visitors = Visitor.query.filter_by(student_id=student.id).order_by(Visitor.visit_date.desc()).all()
    return render_template("my_visitors.html", visitors=visitors)


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        old_pw = request.form.get("old_password", "").strip()
        new_pw = request.form.get("new_password", "").strip()
        confirm_pw = request.form.get("confirm_password", "").strip()
        user = db.session.get(User, session["user_id"])
        if not user or not check_password_hash(user.password, old_pw):
            flash("原密码错误", "danger")
        elif len(new_pw) < 4:
            flash("新密码至少4位", "warning")
        elif new_pw != confirm_pw:
            flash("两次新密码不一致", "warning")
        else:
            user.password = generate_password_hash(new_pw)
            db.session.commit()
            log_operation("修改密码", "user", session["user_id"])
            flash("密码修改成功", "success")
            return redirect(url_for("index"))
    return render_template("change_password.html")


@app.route("/my/profile", methods=["GET", "POST"])
@login_required
def my_profile():
    student = Student.query.filter_by(user_id=session["user_id"]).first()
    if not student:
        flash("未关联学生信息", "warning")
        return redirect(url_for("index"))
    if request.method == "POST":
        student.phone = request.form.get("phone")
        db.session.commit()
        flash("个人信息已更新", "success")
        return redirect(url_for("my_profile"))
    return render_template("my_profile.html", student=student)


# ============================================================
# CLI：初始化数据库
# ============================================================

@app.cli.command("init-db")
def init_db():
    """flask init-db: 创建表并插入示例数据"""
    db.create_all()

    # 宿舍楼
    b1 = Building(name="学生公寓1号楼", floors=6, address="校区A", manager="张阿姨")
    b2 = Building(name="学生公寓2号楼", floors=6, address="校区A", manager="李大叔")
    db.session.add_all([b1, b2])
    db.session.flush()

    # 房间
    rooms = []
    for b in [b1, b2]:
        for floor in range(1, 4):
            for rn in [f"{floor}01", f"{floor}02", f"{floor}03"]:
                rooms.append(Room(
                    room_number=rn,
                    building_id=b.id,
                    capacity=4,
                    occupied=0,
                    room_type="四人间",
                    price=1200.0,
                ))
    db.session.add_all(rooms)
    db.session.flush()

    # 用户（先创建，供学生关联）
    users = [
        User(username="admin", password=generate_password_hash("admin123"), role="admin"),
        User(username="manager", password=generate_password_hash("manager123"), role="dorm_manager", building_id=b1.id),
        User(username="manager2", password=generate_password_hash("manager123"), role="dorm_manager", building_id=b2.id),
        User(username="student1", password=generate_password_hash("123456"), role="student"),
    ]
    db.session.add_all(users)
    db.session.flush()

    # 学生
    students = [
        Student(sno="20230101", name="张三", gender="男", birth="2004-03-12", phone="13800001111", major="计算机科学与技术", class_name="计科2301", user_id=users[3].id),
        Student(sno="20230102", name="李四", gender="女", birth="2004-07-18", phone="13800002222", major="软件工程", class_name="软工2301"),
        Student(sno="20230103", name="王五", gender="男", birth="2004-11-05", phone="13800003333", major="计算机科学与技术", class_name="计科2301"),
        Student(sno="20230104", name="赵六", gender="女", birth="2005-01-20", phone="13800004444", major="数学与应用数学", class_name="数学2301"),
    ]
    db.session.add_all(students)
    db.session.flush()

    # 入住记录
    accs = [
        Accommodation(student_id=students[0].id, room_id=rooms[0].id, check_in_date="2025-09-01", status="入住"),
        Accommodation(student_id=students[1].id, room_id=rooms[3].id, check_in_date="2025-09-01", status="入住"),
        Accommodation(student_id=students[2].id, room_id=rooms[0].id, check_in_date="2025-09-01", status="入住"),
        Accommodation(student_id=students[3].id, room_id=rooms[6].id, check_in_date="2025-09-01", status="入住"),
    ]
    # 更新房间已住人数
    rooms[0].occupied = 2  # 张三 王五
    rooms[3].occupied = 1  # 李四
    rooms[6].occupied = 1  # 赵六
    db.session.add_all(accs)
    db.session.flush()

    # 报修
    repairs = [
        Repair(student_id=students[0].id, room_id=rooms[0].id, description="卫生间水龙头漏水", status="待处理", report_date="2025-12-10"),
        Repair(student_id=students[1].id, room_id=rooms[3].id, description="空调不制冷", status="已完成", report_date="2025-11-20", fix_date="2025-11-22"),
    ]
    db.session.add_all(repairs)

    # 访客
    visitors = [
        Visitor(student_id=students[0].id, visitor_name="张爸爸", id_card="410123198001011234", reason="探望", visit_date="2025-12-01 14:00:00", leave_date="2025-12-01 17:00:00"),
    ]
    db.session.add_all(visitors)

    # 费用
    fees = [
        Fee(student_id=students[0].id, room_id=rooms[0].id, fee_type="住宿费", amount=1200.0, status="已缴", due_date="2025-09-15", pay_date="2025-09-10"),
        Fee(student_id=students[0].id, room_id=rooms[0].id, fee_type="水电费", amount=150.0, status="未缴", due_date="2026-01-10"),
        Fee(student_id=students[1].id, room_id=rooms[3].id, fee_type="住宿费", amount=1200.0, status="未缴", due_date="2025-09-15"),
    ]
    db.session.add_all(fees)

    # 公告
    db.session.add(Announcement(
        title="欢迎使用学生公寓管理系统",
        content="系统已正式上线，请同学们及时查看费用信息，如有报修需求请通过系统提交。",
        user_id=users[0].id,
    ))

    db.session.commit()
    print("数据库初始化完成！")


@app.cli.command("reset-db")
def reset_db():
    """flask reset-db: 删除所有表并重新初始化"""
    db.drop_all()
    db.create_all()
    print("数据库已重置，正在导入种子数据...")
    init_db()


@app.cli.command("init-advanced")
def init_advanced():
    """flask init-advanced: 导入存储过程/函数/触发器"""
    import os as _os
    sql_path = _os.path.join(_os.path.dirname(__file__), "db_advanced.sql")
    with open(sql_path, "r", encoding="utf-8") as f:
        sql = f.read()
    # 按 DELIMITER 分句执行
    statements = sql.replace("DELIMITER $$", "").replace("DELIMITER ;", "").split("$$")
    for stmt in statements:
        stmt = stmt.strip()
        if stmt:
            try:
                db.session.execute(text(stmt))
            except Exception as e:
                print(f"跳过（可能已存在）: {str(e)[:80]}")
    db.session.commit()
    print("高级特性初始化完成！")


if __name__ == "__main__":
    app.run(debug=True, host='0.0.0.0', port=5000, use_reloader=False)
